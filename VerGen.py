#To run the file streamlit run <fileName.py>

import pandas as pd
import nltk
import re
import os
from nltk.tokenize import sent_tokenize
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter
import streamlit as st
import matplotlib.pyplot as plt
import seaborn as sns


# Download necessary NLTK resources
nltk.download("punkt")

# Function to clean pattern
def clean_heading(text):
    """Extract only the heading from a sentence."""
    text = text.strip()
    text = re.sub(r"^[\s]*[\u2022\u25E6\u2023⦁•\d]+\.*\)*\s*", "", text)
    text = re.sub(r"\s*:\s*$", "", text)
    text = re.sub(r"[^a-zA-Z0-9\s\-\:]", "", text)
    text = text.lower().strip()
    return text

# Function to check pattern
def check_pattern(text):
    if pd.isna(text):
        return pd.Series(["No", "No It is nan", ""])
    text = str(text).strip()
    required_ruleBook = ["acceptance criteria", "input", "output", "pre-condition"]
    lines = text.split("\n")
    sentences = []
    for line in lines:
        line = line.strip()
        first_word = re.sub(r"^[\s]*[\u2022\u25E6\u2023⦁•\d]+\.*\)*\s*", "", line.split(":")[0].strip().lower())
        if first_word in required_ruleBook:
            sentences.append(first_word)
        else:
            sentences.extend(sent_tokenize(line))
    found_ruleBook = {}
    for sent in sentences:
        clean_text = clean_heading(sent)
        if clean_text in required_ruleBook:
            found_ruleBook[clean_text] = clean_text
    missing = [h for h in required_ruleBook if h not in found_ruleBook]
    fixed_suggested_pattern = """Pre-Condition:\nAcceptance Criteria:\n Input:\n Output:"""
    validation_status = "Matched with RuleBook" if not missing else "Not Matched with RuleBook"
    missing_info = f"Missing: {', '.join(missing)}" if missing else ""
    suggested_pattern = "" if validation_status == "Matched with RuleBook" else fixed_suggested_pattern
    return pd.Series([validation_status, missing_info, suggested_pattern])

# Streamlit Application
def main():
    st.title("Verification Criteria Validation Dashboard")
    
    uploaded_file = st.file_uploader("Choose an Excel file", type="xlsx")
    
    if uploaded_file is not None:
        df = pd.read_excel(uploaded_file)
        
        if "DA_Verification_Criteria" not in df.columns:
            st.error("The uploaded file must contain a column named 'DA_Verification_Criteria'.")
            return

        df["DA_Verification_Criteria"] = df["DA_Verification_Criteria"].astype(str)
        new_columns = df["DA_Verification_Criteria"].apply(check_pattern)
        new_columns.columns = ["Verification Criteria Validation Status", "Missing Rule Patterns", "Suggested Rule Book Pattern"]
        
        insert_index = df.columns.get_loc("DA_Verification_Criteria") + 1
        df = pd.concat([df.iloc[:, :insert_index], new_columns, df.iloc[:, insert_index:]], axis=1)

        # Get user input for the file name
        output_file_name = st.text_input("Enter the name of the output file", "processed_output.xlsx")
        
        # Save the output file with colors and wrap text
        with pd.ExcelWriter(output_file_name, engine="openpyxl") as writer:
            df.to_excel(writer, index=False)
            workbook = writer.book
            worksheet = writer.sheets["Sheet1"]
            fill_green = PatternFill(start_color="006400", end_color="006400", fill_type="solid")
            fill_red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            wrap_text = Alignment(wrap_text=True)
            
            validated_col_idx = df.columns.get_loc("Verification Criteria Validation Status") + 1
            for row in range(2, len(df) + 2):
                cell = worksheet.cell(row=row, column=validated_col_idx)
                if cell.value == "Matched with RuleBook":
                    cell.fill = fill_green
                elif cell.value == "Not Matched with RuleBook":
                    cell.fill = fill_red
            
            for col_idx, column in enumerate(df.columns, 1):
                col_letter = get_column_letter(col_idx)
                for row_idx, value in enumerate(df[column], 2):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
                    max_length = 0
                    for row in df[column]:
                        try:
                            if len(str(row)) > max_length:
                                max_length = len(str(row))
                        except:
                            pass
                    worksheet.column_dimensions[col_letter].width = 50

            workbook.save(output_file_name)

        # Provide download button for the processed file
        with open(output_file_name, "rb") as file:
            st.download_button(
                label="Download Processed File",
                data=file.read(),
                file_name=output_file_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        
        # Pie chart visualization
        yes_count = df["Verification Criteria Validation Status"].value_counts().get("Matched with RuleBook", 0)
        no_count = df["Verification Criteria Validation Status"].value_counts().get("Not Matched with RuleBook", 0)
        total_count = len(df)

        st.subheader("Validation Status Summary")
        fig_pie, ax_pie = plt.subplots(figsize=(6, 6))
        labels = ['Matched with RuleBook', 'Not Matched with RuleBook']
        sizes = [yes_count, no_count]
        colors = ['green', 'red']
        explode = (0.1, 0)  # Highlight "Yes"
        ax_pie.pie(sizes, labels=labels, autopct="%1.1f%%", colors=colors, explode=explode, shadow=True, startangle=140)
        ax_pie.set_title(f"Validation Summary (Total: {total_count})")
        st.pyplot(fig_pie)

        # 📌 Bar Chart
        fig_bar, ax_bar = plt.subplots(figsize=(6, 4))
        ax_bar.bar(labels, sizes, color=colors)
        ax_bar.set_title("Validation Count")
        ax_bar.set_ylabel("Count")
        ax_bar.set_xlabel("Validation Status")
        st.pyplot(fig_bar)
       # plt.pie(sizes, labels=labels, colors=colors, autopct=lambda pct: func(pct, sizes), startangle=90)
        plt.axis('equal')

        plt.title("Verification Validation Status")
        #st.pyplot(plt)

if __name__ == "__main__":
    main()
